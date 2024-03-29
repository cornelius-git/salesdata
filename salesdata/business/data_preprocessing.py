import pandas as pd
import numpy as np
from sql_setting import *
from openpyxl import load_workbook
import openpyxl

def orignal_import(file_path="D:\\1何军\\财务系统\\系统导出数据",log_name="log.csv",del_name="del.csv"):
    # 系统原始数据导入
    log_table = pd.read_csv(r"{}\{}".format(file_path,log_name))
    log_table.to_sql("user_account_integral_log", engine, if_exists="replace", index=False)
    del_table = pd.read_csv(r"{}\{}".format(file_path, del_name))
    del_table.to_sql("user_account_integral_del", engine, if_exists="replace", index=False)
    sql = 'UPDATE user_account_integral_del SET action_id = "107" ,action_name = "EVOS购车客户定向权益" ' \
          'WHERE user_id = 66670 and action_id = 112'
    cursor.execute(sql)
    cursor.commit()
    print("数据导入完成")




def jingdong_import(file_path,jingdong_name,sheet_name):
    # 京东对账单原始数据导入
    jingdong = pd.read_excel(r"{}\{}".format(file_path,jingdong_name),sheet_name=sheet_name)
    if "退款金额" in jingdong.columns:
        jingdong["实际支付金额"] = jingdong["商品含税总额"] - jingdong["退款金额"]
    else:
        jingdong["实际支付金额"] = jingdong["商品含税总额"]
    filter_str = ['业务单号(子订单号)', '原始订单号', '下单时间', '商品类型', '商品编号', '商品名称', '商品税率', '商品数量', '商品含税单价',
                         '商品含税总额',  "实际支付金额"]
    rename_dict = {"业务单号(子订单号)": "child_number", "原始订单号": "origin_number", "下单时间": "order_time",
                             "商品类型": "business_type", "商品编号": "business_code",
                             "商品名称": "business_name", "商品税率": "business_tax", "商品数量": "jbussiness_number",
                             "商品含税单价": "bussiness_tax_price", "商品含税总额": "business_total_tax_price",
                              "实际支付金额": "pay_price"}
    if "月份" in jingdong.columns.tolist():
        filter_str.append("月份")
        rename_dict["月份"] = "filter_month"
    jingdong = jingdong[filter_str]
    # '商品未税总额（未税单价*数量）', '商品未税单价',
    jingdong.rename(columns=rename_dict, inplace=True)
    jingdong["child_number_code"] = jingdong.apply(lambda x: "jd" + str(x["child_number"])+str(x["business_code"]),axis=1)
    # business = jingdong[jingdong["business_type"] != "运费"]
    jingdong.reset_index(drop=True,inplace=True)
    jingdong.to_sql("jingdong", engine, if_exists="replace")
    # "商品未税总额（未税单价*数量）": "business_total_price",
    #                              "商品未税单价": "bussiness_price",
    # 运费分摊到每个商品上面

    # jingdong[["origin_number", "business_code"]] = jingdong[["origin_number", "business_code"]].astype(str)
    # jingdong["business_code"] = jingdong["business_code"].apply(lambda x: str(x).replace("fre", ""))
    # jingdong["Identification_code"] = jingdong["origin_number"] + jingdong["business_code"]
    #
    # freight = jingdong[jingdong["business_type"] == "运费"][["Identification_code","origin_number", "business_code",
    #                                                        "bussiness_tax_price"]]
    #
    # freight.drop(["origin_number","business_code"],axis=1,inplace=True)
    # freight.rename(columns={"bussiness_tax_price":"flow_fee"},inplace=True)
    # business = jingdong[jingdong["business_type"] != "运费"]
    # 原始运费重构
    # count_number = pd.pivot_table(business, index="origin_number", values="bussiness_number", aggfunc=np.sum)
    #
    # count_number.reset_index(inplace=True)

    # freight = pd.pivot_table(freight,index="origin_number", values="business_total_tax_price", aggfunc=np.sum)
    #
    #
    # freight = pd.merge(count_number, freight, how="left", left_on="origin_number",right_on="origin_number")
    #
    # freight.columns = ["origin_number", "bussiness_total_number", "flow_fee"]
    # freight.fillna(0, inplace=True)
    # freight.to_excel("yunfei.xlsx")
    # business.to_excel("diyi.xlsx")
    # result = pd.merge(business, freight, how="left", on="Identification_code")
    # result.fillna(0,inplace=True)
    # result["real_pay_price"] = result["flow_fee"] / result["bussiness_total_number"] * result["bussiness_number"] + \
    #                            result["pay_price"]

    # result["real_pay_price"] = result.loc[:,"bussiness_tax_price"] + result.loc[:,"flow_fee"]
    #
    # result.to_sql("jingdong", engine, if_exists="replace")


def jingdong_relationship(file_path,file_name,sheet_list):
    # 京东对应关系表导入
    i=0
    for sheet in sheet_list:
        relation_file = pd.read_excel(r"{}\{}".format(file_path, file_name),sheet_name=sheet)
        relation_file = relation_file[['京东订单号', '订单号', '电话', '商品编号','商品名', '商城单价', '数量', '商品品类']]
        # '物流单号',
        relation_file.rename(columns={'京东订单号': 'child_number',
                                     '订单号': 'business_id', '电话': 'oder_phone', '商品编号':'business_code','商品名': 'business_name',
                                     '商城单价': 'business_price', '数量': 'business_number', '商品品类': 'business_type'},
                            inplace=True)
        relation_file =  relation_file.astype(str)
        relation_file["business_id_code"] = relation_file.apply(lambda x: str(x['business_id'])+str(x['business_code']).replace(" ",""),axis=1)
        relation_file['child_number_code'] = relation_file.apply(lambda x: "jd"+str(x['child_number'])+str(x['business_code']).replace(" ",""),axis=1)
        # '物流单号': ' logistics_number',
        if i == 0:
            relation_file.to_sql("guanxi", engine, if_exists="replace",index=False)
            i+=4
        else:
            relation_file.to_sql("guanxi", engine, if_exists="append", index=False)

def action_relationship(file_path,file_name,sheet_name):
    # pid关系导入 禁用
    pid = pd.read_excel(r"{}\{}".format(file_path,file_name), sheet_name=sheet_name)
    pid = pid[['对应活动action_id', 'program_id', '对应活动action_name', '有偿无偿']]
    pid.rename(columns={'对应活动action_id': 'action_id', 'program_id': 'program_id', '对应活动action_name': 'action_name',
                        '有偿无偿': 'judge'}, inplace=True)
    pid.to_sql("action_programid",engine, if_exists="replace", index=False)

def pid_aid(file_path="D:\\1何军\\财务系统\\系统导出数据",file_name="Program_ID与Action_ID对应关系表导入版本.xlsx"):
    # pid和action_id的对应关系导入

    sql = "TRUNCATE action_program"
    cursor.execute(sql)
    cursor.commit()
    wt = load_workbook(filename=r"{}\{}".format(file_path,file_name))
    common_2021 = wt["2021年"]
    common_2022 = wt["2022常规行为"]
    active_2022 = wt["2022其他行为"]
    for i in range(2,common_2021.max_row):
        if common_2021.cell(row=i,column=1).value:
            sql = "INSERT INTO action_program(action_id,action_name," \
                  "program_id,integral_judge,action_type,use_type,start_month,end_month) VALUES ('{}','{}','{}','{}','{}','{}','{}','{}') "\
                .format(common_2021.cell(row=i,column=1).value,common_2021.cell(row=i,column=3).value,
                        common_2021.cell(row=i,column=2).value,common_2021.cell(row=i,column=4).value,
                        common_2021.cell(row=i,column=5).value,common_2021.cell(row=i,column=6).value,
                        "2021-01","2021-12")
            cursor.execute(sql)
        else:
            break
    cursor.commit()
    for i in range(2,common_2022.max_row):
        if common_2022.cell(row=i,column=1).value:
            sql = "INSERT INTO action_program(action_id,action_name," \
                  "program_id,integral_judge,action_type,use_type,start_month,end_month) VALUES ('{}','{}','{}','{}','{}','{}','{}','{}') "\
                .format(common_2022.cell(row=i,column=1).value,common_2022.cell(row=i,column=2).value,
                        common_2022.cell(row=i,column=6).value,common_2022.cell(row=i,column=3).value,
                        common_2022.cell(row=i,column=4).value,common_2022.cell(row=i,column=5).value,
                        "2022-01","2022-03")
            cursor.execute(sql)
            sql1 = "INSERT INTO action_program(action_id,action_name," \
                  "program_id,integral_judge,action_type,use_type,start_month,end_month) VALUES ('{}','{}','{}','{}','{}','{}','{}','{}') " \
                .format(common_2022.cell(row=i, column=1).value, common_2022.cell(row=i, column=2).value,
                        common_2022.cell(row=i, column=7).value, common_2022.cell(row=i, column=3).value,
                        common_2022.cell(row=i, column=4).value, common_2022.cell(row=i, column=5).value,
                        "2022-04", "2022-06")
            cursor.execute(sql1)
            sql1 = "INSERT INTO action_program(action_id,action_name," \
                   "program_id,integral_judge,action_type,use_type,start_month,end_month) VALUES ('{}','{}','{}','{}','{}','{}','{}','{}') " \
                .format(common_2022.cell(row=i, column=1).value, common_2022.cell(row=i, column=2).value,
                        common_2022.cell(row=i, column=8).value, common_2022.cell(row=i, column=3).value,
                        common_2022.cell(row=i, column=4).value, common_2022.cell(row=i, column=5).value,
                        "2022-07", "2022-09")
            cursor.execute(sql1)
        else:
            break
    cursor.commit()
    for i in range(2,active_2022.max_row):
        if active_2022.cell(row=i, column=1).value:
            sql = "INSERT INTO action_program(action_id,action_name," \
                  "program_id,integral_judge,action_type,use_type,start_month,end_month) VALUES ('{}','{}','{}','{}','{}','{}','{}','{}') " \
                .format(active_2022.cell(row=i, column=1).value, active_2022.cell(row=i, column=3).value,
                        active_2022.cell(row=i, column=2).value, active_2022.cell(row=i, column=4).value,
                        active_2022.cell(row=i, column=5).value, active_2022.cell(row=i, column=6).value,
                        "2022-01", "2023-04")
            print(active_2022.cell(row=i, column=1).value)
            cursor.execute(sql)
        else:
            break
    cursor.commit()

def fuyu_jingdong(file_path="D:\\1何军\\财务系统\\系统导出数据",file_name="商城订单1.xlsx"):
    # 福域订单导入
    fuyu = pd.read_excel(r"{}\{}".format(file_path,file_name),converters={"实际单价": int,"数量":int,"来源spu":str})
    # fuyu = fuyu[(fuyu["订单状态"] != 7) & (fuyu["退款状态"] != 1) & (fuyu["订单状态"] != 4)]
    fuyu = fuyu[fuyu["退款状态"] != 1]
    fuyu = fuyu[["订单号","供货商","SKU","数量","来源spu","实际单价","商品名称","订单状态"]]
    fuyu.rename(columns={"订单号":"business_id","供货商":"business_channel","SKU":"sku",
                         "数量":"business_quantity","来源spu":"spu","实际单价":"business_price",
                         "订单状态":"order_statue","商品名称":"business_name"},inplace=True)
    # print(fuyu.info)
    # fuyu["business_single_price"] = fuyu["business_price"]-fuyu["business_quantity"]

    fuyu.reset_index(drop=True,inplace=True)
    fuyu["business_id_code"] = fuyu.apply(lambda x: str(x["business_id"])+str(x["spu"]),axis=1)

    # fuyu["business_id"] = fuyu["business_id"].apply(lambda x: str(x).split("_")[0])

    fuyu.to_sql("business_relationship",engine,if_exists="replace")
    sql = 'UPDATE business_relationship set spu="5237209" WHERE business_id="M0577127583677861888" and spu = "5158704"'
    cursor.execute(sql)
    cursor.commit()

def fuyu_skuinfo(file_path="D:\\1何军\\财务系统\\系统导出数据",file_name="商城订单sku.csv"):
    fuyu = pd.read_csv(r"{}\{}".format(file_path, file_name))
    fuyu = fuyu[fuyu["refund_stauts"] != 1]
    # fuyu = fuyu[fuyu["send_num"].notnull()]
    fuyu = fuyu[["order_no", "buy_num", "sku_code", "shared_fb", "spu_name", "business_channel"]]
    fuyu.rename(columns={"order_no": "business_id", "business_channel": "business_channel", "sku_code": "sku",
                         "buy_num": "business_quantity",  "shared_fb": "business_price",
                         "spu_name": "business_name"}, inplace=True)

    fuyu.reset_index(drop=True, inplace=True)

    # fuyu["business_id"] = fuyu["business_id"].apply(lambda x: str(x).split("_")[0])
    fuyu.to_sql("business_relationship", engine, if_exists="replace")

def caiwu_account(file_path="D:\\1何军\\财务系统",file_name="积分原始数据.xlsx"):
    # 积分原始对账数据
    caiwu = pd.read_excel(r"{}\{}".format(file_path,file_name),dtype={"sale_month": str})
    six_caiwu = pd.read_excel(r"{}\{}".format(file_path,"6原始数据.xlsx"),dtype={"sale_month": str})
    seven_caiwu = pd.read_excel(r"{}\{}".format(file_path, "7原始数据.xlsx"), dtype={"sale_month": str})
    sum_caiwu = pd.concat([caiwu,six_caiwu,seven_caiwu],axis=0)
    sum_caiwu["sale_month"] = sum_caiwu["sale_month"].apply(lambda x: str(x)[:7])
    sum_caiwu.reset_index(drop=True,inplace=True)
    sum_caiwu.to_sql("caiwu_account",engine, if_exists="replace")

def dashang_import():
    sibada = pd.read_excel(r"D:\1何军\财务系统\系统导出数据\dashang.xlsx")
    sibada.to_sql("dashang", engine, if_exists="replace", index=False)

if __name__=="__main__":
    # 原始数据导入
    # orignal_import()
    # fuyu_jingdong()
    # fuyu_skuinfo()
    # action_id 和progra_id对应关系导入
    # pid_aid()
    # caiwu_account()
    #     打赏
    # dashang_import()
    # fuyu_jingdong()
    #京东原始数据导入
    jingdong_import(file_path="D:\\1何军\\财务对账\\4月",jingdong_name="京东原始数据0325-0424.xlsx",sheet_name="Sheet1")
    # 闫浩的关系导入
    # jingdong_relationship(file_path="D:\\1何军\\财务系统",file_name="福域积分商城订单明细表（1月-6月）.xlsx",
    #                       sheet_list=["1月","2月","3月","4月","5月","6月"])
    # jingdong_import(file_path="D:\\1何军\\京东数据\\2报账单",jingdong_name="1-5月原始数据汇总.xlsx",sheet_name="Sheet1")



